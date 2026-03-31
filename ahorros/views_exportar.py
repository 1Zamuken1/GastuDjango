"""
Vistas de exportación de metas de ahorro.
Soporta CSV, Excel (openpyxl) y PDF (reportlab).

Parámetros GET:
- estado : Todos los estados (vacío), SIN_INICIAR, ACTIVO, COMPLETADO, ABANDONADO
"""
import csv
import io
from datetime import date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone

from .models import AhorroMeta


def _build_qs(request):
    """
    Construye el queryset de ahorros filtrado según los parámetros GET.
    Siempre filtra por request.user.
    """
    estado = request.GET.get('estado', '')

    qs = AhorroMeta.objects.filter(
        usuario=request.user
    ).select_related('categoria').order_by('-fecha_creacion')

    if estado in dict(AhorroMeta.Estado.choices):
        qs = qs.filter(estado=estado)

    return qs, estado


def _nombre_archivo(estado, extension):
    """Genera el nombre del archivo de descarga."""
    estado_str = estado.lower() if estado else 'todas'
    hoy = timezone.now().strftime("%Y-%m-%d")
    return f"gastuapp_ahorros_{estado_str}_{hoy}.{extension}"


@login_required
def exportar_csv(request):
    """Exporta metas de ahorro a CSV."""
    qs, estado = _build_qs(request)
    nombre = _nombre_archivo(estado, 'csv')

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'

    writer = csv.writer(response)
    writer.writerow(['Categoría', 'Descripción', 'Estado', 'Progreso', 'Total Acumulado', 'Monto Meta', 'Fecha Límite', 'Frecuencia', 'Cuotas'])

    for a in qs:
        pct = (a.total_acumulado / a.monto_meta * 100) if (a.total_acumulado and a.monto_meta) else 0
        writer.writerow([
            a.categoria.nombre if a.categoria else 'N/A',
            a.descripcion or '',
            a.get_estado_display(),
            f"{pct:.1f}%",
            f"${a.total_acumulado:.2f}" if a.total_acumulado else "$0.00",
            f"${a.monto_meta:.2f}" if a.monto_meta else "$0.00",
            a.fecha_meta.strftime('%d/%m/%Y') if a.fecha_meta else 'N/A',
            a.get_frecuencia_display(),
            a.cantidad_cuotas,
        ])

    return response


@login_required
def exportar_excel(request):
    """Exporta metas de ahorro a Excel usando openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl no está instalado.', status=500)

    qs, estado = _build_qs(request)
    nombre = _nombre_archivo(estado, 'xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ahorros'

    COLOR_HEADER  = '0F172A'
    COLOR_ROW     = 'FFFBEB'
    
    thin = Side(style='thin', color='E2E8F0')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:H1')
    titulo = ws['A1']
    estado_lbl = f" - Estado: {estado}" if estado else ""
    titulo.value = f'GastuApp — Mis Metas de Ahorro{estado_lbl}'
    titulo.font      = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    titulo.fill      = PatternFill('solid', fgColor=COLOR_HEADER)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    encabezados = ['Categoría', 'Descripción', 'Estado', 'Progreso', 'Acumulado', 'Monto Meta', 'Límite', 'Frecuencia']
    for col, tit in enumerate(encabezados, 1):
        celda = ws.cell(row=2, column=col, value=tit)
        celda.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        celda.fill      = PatternFill('solid', fgColor='D97706') # Amber 600
        celda.alignment = Alignment(horizontal='center', vertical='center')
        celda.border    = borde
    ws.row_dimensions[2].height = 20

    for fila, a in enumerate(qs, 3):
        pct = (a.total_acumulado / a.monto_meta * 100) if (a.total_acumulado and a.monto_meta) else 0
        valores = [
            a.categoria.nombre if a.categoria else 'N/A',
            a.descripcion or '',
            a.get_estado_display(),
            f"{pct:.1f}%",
            float(a.total_acumulado or 0),
            float(a.monto_meta or 0),
            a.fecha_meta.strftime('%d/%m/%Y') if a.fecha_meta else 'N/A',
            a.get_frecuencia_display(),
        ]
        for col, val in enumerate(valores, 1):
            celda = ws.cell(row=fila, column=col, value=val)
            celda.fill   = PatternFill('solid', fgColor=COLOR_ROW)
            celda.border = borde
            celda.font   = Font(name='Calibri', size=10)
            if col in (5, 6):
                celda.number_format = '#,##0.00'
                celda.alignment     = Alignment(horizontal='right')
            elif col == 4:
                celda.alignment     = Alignment(horizontal='right')

    anchos = [20, 30, 15, 12, 16, 16, 15, 15]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


@login_required
def exportar_pdf(request):
    """Exporta metas de ahorro a PDF usando reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse('reportlab no está instalado.', status=500)

    qs, estado = _build_qs(request)
    nombre = _nombre_archivo(estado, 'pdf')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    COLOR_HEADER = colors.HexColor('#0F172A')
    COLOR_ROW    = colors.HexColor('#FFFBEB')
    
    elementos = []

    estado_lbl = f" - Estado: {estado}" if estado else ""
    header_data = [[
        Paragraph(f'<font color="white"><b>GastuApp — Mis Metas de Ahorro{estado_lbl}</b></font>', styles['Normal']),
        Paragraph(f'<font color="#94a3b8">Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}</font>', styles['Normal']),
    ]]
    header_table = Table(header_data, colWidths=['60%', '40%'])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), COLOR_HEADER),
        ('PADDING',    (0,0), (-1,-1), 10),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TEXTCOLOR',  (1,0), (1,0),   colors.HexColor('#94a3b8')),
        ('ALIGN',      (1,0), (1,0),   'RIGHT'),
    ]))
    elementos.append(header_table)
    elementos.append(Spacer(1, 0.4*cm))

    encabezados = ['Categoría', 'Descripción', 'Estado', 'Progreso', 'Acumulado', 'Meta', 'Límite']
    filas = [encabezados]
    for a in qs:
        pct = (a.total_acumulado / a.monto_meta * 100) if (a.total_acumulado and a.monto_meta) else 0
        filas.append([
            a.categoria.nombre if a.categoria else 'N/A',
            a.descripcion or '—',
            a.get_estado_display(),
            f"{pct:.1f}%",
            f'${a.total_acumulado:,.2f}' if a.total_acumulado else '$0.00',
            f'${a.monto_meta:,.2f}' if a.monto_meta else '$0.00',
            a.fecha_meta.strftime('%d/%m/%Y') if a.fecha_meta else 'N/A',
        ])

    tabla = Table(filas, colWidths=[4*cm, 7.5*cm, 2.8*cm, 2.2*cm, 3.5*cm, 3.5*cm, 3*cm], repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0),  colors.HexColor('#D97706')), # Amber 600
        ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
        ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 9),
        ('ALIGN',       (3,0), (5,-1),  'RIGHT'),
        ('GRID',        (0,0), (-1,-1), 0.4, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [COLOR_ROW, colors.white]),
        ('PADDING',     (0,0), (-1,-1), 6),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response
