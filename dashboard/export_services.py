"""
Servicios de exportacion del dashboard.

Genera reportes Excel y PDF del resumen mensual completo
(movimientos + aportes de ahorro). Usado por las vistas
de exportacion del dashboard.
"""
import io
import os
from datetime import date

from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string

from gastu_django.constants import MESES_ES


# ── Excel ────────────────────────────────────────────────────────────────────

def generar_excel_dashboard(ctx, all_items, mes, anio):
    """
    Genera un archivo Excel con el reporte mensual del dashboard.

    Args:
        ctx (dict): contexto del dashboard (build_dashboard_context).
        all_items (list[dict]): items normalizados (obtener_items_completos_mes).
        mes (int): mes del reporte.
        anio (int): ano del reporte.

    Returns:
        HttpResponse: archivo Excel para descarga.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f'{MESES_ES[mes]} {anio}'

    # Estilos
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Titulo
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f'Gastu \u2014 Reporte {MESES_ES[mes]} {anio}'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='0F172A')
    title_cell.alignment = Alignment(horizontal='center')

    # Resumen
    ws['A3'] = 'Ingresos'
    ws['B3'] = float(ctx['total_ingresos'])
    ws['B3'].number_format = '$#,##0'
    ws['C3'] = 'Egresos'
    ws['D3'] = float(ctx['total_egresos'])
    ws['D3'].number_format = '$#,##0'
    ws['A4'] = 'Ahorros'
    ws['B4'] = float(ctx['ahorros_mes'])
    ws['B4'].number_format = '$#,##0'
    ws['C4'] = 'Utilidad'
    ws['D4'] = float(ctx['utilidad'])
    ws['D4'].number_format = '$#,##0'
    for cell in ['A3', 'C3', 'A4', 'C4']:
        ws[cell].font = Font(bold=True, size=10, color='64748B')

    # Cabeceras de tabla
    headers = ['Tipo', 'Descripcion', 'Categoria', 'Fecha', 'Monto']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Datos
    for i, item in enumerate(all_items, 7):
        ws.cell(row=i, column=1, value=item['tipo']).border = thin_border
        ws.cell(row=i, column=2, value=item['descripcion']).border = thin_border
        ws.cell(row=i, column=3, value=item['categoria']).border = thin_border
        ws.cell(row=i, column=4, value=item['fecha']).border = thin_border
        c = ws.cell(row=i, column=5, value=item['monto'])
        c.number_format = '$#,##0'
        c.border = thin_border
        # Color por tipo
        if item['tipo'] == 'INGRESO':
            c.font = Font(color='059669', bold=True)
        elif item['tipo'] == 'EGRESO':
            c.font = Font(color='E11D48', bold=True)
        else:
            c.font = Font(color='D97706', bold=True)

    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Gastu_Reporte_{MESES_ES[mes]}_{anio}.xlsx"'
    wb.save(response)
    return response


# ── PDF ──────────────────────────────────────────────────────────────────────

def generar_pdf_dashboard(ctx, all_items, mes, anio, user):
    """
    Genera un archivo PDF con el reporte mensual del dashboard.

    Args:
        ctx (dict): contexto del dashboard (build_dashboard_context).
        all_items (list[dict]): items normalizados (obtener_items_completos_mes).
        mes (int): mes del reporte.
        anio (int): ano del reporte.
        user: instancia del usuario.

    Returns:
        HttpResponse: archivo PDF para descarga.
    """
    from xhtml2pdf import pisa

    hoy = date.today()
    logo_path = os.path.join(
        settings.BASE_DIR, 'static', 'img', 'gastu_grafica_pdf.jpg'
    ).replace('\\', '/')

    pdf_ctx = {
        'mes_nombre':       MESES_ES[mes],
        'anio':             anio,
        'total_ingresos':   float(ctx['total_ingresos']),
        'total_egresos':    float(ctx['total_egresos']),
        'ahorros_mes':      float(ctx['ahorros_mes']),
        'utilidad':         float(ctx['utilidad']),
        'disponible':       float(ctx['disponible']),
        'movimientos':      all_items,
        'usuario':          user.username,
        'fecha_generacion': hoy.strftime('%d/%m/%Y %H:%M'),
        'logo_path':        logo_path,
    }

    html_string = render_to_string('dashboard/reporte_pdf.html', pdf_ctx)

    buffer = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=buffer, encoding='utf-8')

    if pisa_status.err:
        return HttpResponse('Error generando el PDF', status=500)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Gastu_Reporte_{MESES_ES[mes]}_{anio}.pdf"'
    return response
