import json
import calendar
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.db.models.functions import TruncDate
from django.http import JsonResponse
from django.shortcuts import render

from dashboard.models import ResumenMensual
from movimientos.models import Movimiento
from notificaciones.models import Notificacion
from ahorros.models import AporteAhorro, AhorroMeta


MESES_ES = {
    1: 'Enero',   2: 'Febrero',  3: 'Marzo',    4: 'Abril',
    5: 'Mayo',    6: 'Junio',    7: 'Julio',     8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}

ZERO = Decimal('0')

PIE_COLORES = [
    '#e11d48', '#f87171', '#fbbf24', '#a3e635',
    '#34d399', '#38bdf8', '#818cf8', '#f472b6',
]


def _ultimo_dia(mes, anio):
    """Devuelve el último día del mes dado."""
    return date(anio, mes, calendar.monthrange(anio, mes)[1])


def _totales_movimiento(user, mes, anio):
    """Calcula ingresos y egresos directamente desde Movimiento (fallback sin ResumenMensual)."""
    qs = Movimiento.objects.filter(
        usuario=user, activo=True,
        fecha_registro__month=mes,
        fecha_registro__year=anio,
    )
    ingresos = qs.filter(tipo='INGRESO').aggregate(t=Sum('monto'))['t'] or ZERO
    egresos  = qs.filter(tipo='EGRESO').aggregate(t=Sum('monto'))['t'] or ZERO
    return ingresos, egresos


def _build_context(user, mes, anio):
    """
    Construye todos los datos del dashboard para un mes/año dado.
    Fuente principal: ResumenMensual (Opcion A — lectura directa, sin recalculo).
    Ahorros se calculan directamente desde AporteAhorro porque ResumenMensual
    no los integra todavia.
    """
    hoy           = date.today()
    es_mes_actual = (mes == hoy.month and anio == hoy.year)
    ultimo_dia    = hoy if es_mes_actual else _ultimo_dia(mes, anio)

    resumen = ResumenMensual.objects.filter(
        usuario=user, mes=mes, anio=anio,
    ).first()

    if resumen:
        # Leer totales directamente del cache pre-calculado por signals.
        # No llamar a _totales_movimiento() — es redundante y costoso.
        total_ingresos = resumen.total_ingresos
        total_egresos  = resumen.total_egresos
        total_ahorros  = resumen.total_ahorros
        utilidad       = resumen.ingreso_neto
        disponible     = resumen.ganancia_acumulada
        hay_deficit    = resumen.deficit
    else:
        # Fallback: usuario sin historial o primer movimiento del mes.
        total_ingresos, total_egresos = _totales_movimiento(user, mes, anio)
        total_ahorros = ZERO
        utilidad      = total_ingresos - total_egresos
        disponible    = utilidad
        hay_deficit   = total_egresos > total_ingresos

    diferencia = total_ingresos - total_egresos

    # Ahorros del mes (query directa — ResumenMensual.total_ahorros es siempre 0
    # hasta que el modulo ahorros conecte sus propios signals)
    ahorros_mes = (
        AporteAhorro.objects
        .filter(
            ahorro__usuario=user,
            fecha_registro__month=mes,
            fecha_registro__year=anio,
        )
        .aggregate(t=Sum('aporte'))['t'] or ZERO
    )

    # Ahorro total acumulado hasta el ultimo dia del mes visto
    ahorro_total = (
        AporteAhorro.objects
        .filter(
            ahorro__usuario=user,
            fecha_registro__lte=ultimo_dia,
        )
        .aggregate(t=Sum('aporte'))['t'] or ZERO
    )

    # Pie chart — egresos por categoria
    egresos_cat = (
        Movimiento.objects
        .filter(
            usuario=user, tipo='EGRESO', activo=True,
            fecha_registro__month=mes,
            fecha_registro__year=anio,
        )
        .values('categoria__nombre')
        .annotate(total=Sum('monto'))
        .order_by('-total')[:8]
    )

    pie_labels  = [item['categoria__nombre'] or 'Sin categoria' for item in egresos_cat]
    pie_valores = [float(item['total']) for item in egresos_cat]
    pie_data = {
        'labels':  pie_labels,
        'valores': pie_valores,
        'colores': PIE_COLORES[:len(pie_labels)],
    }

    # Metas de ahorro activas — reemplaza Top Gastos (redundante con pie)
    metas_raw = (
        AhorroMeta.objects
        .filter(usuario=user, estado='ACTIVO')
        .select_related('categoria')
        .order_by('-fecha_creacion')[:5]
    )
    metas_ahorro_activas = []
    for m in metas_raw:
        meta_monto = float(m.monto_meta) if m.monto_meta else 1
        acumulado  = float(m.total_acumulado) if m.total_acumulado else 0
        pct = min(round(acumulado / meta_monto * 100, 1), 100) if meta_monto > 0 else 0
        metas_ahorro_activas.append({
            'descripcion':  m.descripcion or m.categoria.nombre,
            'categoria':    m.categoria.nombre if m.categoria else 'Sin categoría',
            'acumulado':    acumulado,
            'acumulado_fmt': f"${acumulado:,.0f}",
            'meta':         meta_monto,
            'meta_fmt':     f"${meta_monto:,.0f}",
            'pct':          pct,
            'frecuencia':   m.get_frecuencia_display(),
        })

    # Movimientos unificados: Ingresos + Egresos + Ahorros
    movs_raw = list(
        Movimiento.objects
        .filter(
            usuario=user, activo=True,
            fecha_registro__month=mes,
            fecha_registro__year=anio,
        )
        .select_related('categoria')
        .order_by('-fecha_registro')[:15]
    )
    ahorros_raw = list(
        AporteAhorro.objects
        .filter(
            ahorro__usuario=user,
            estado_ap='APORTADO',
            fecha_registro__month=mes,
            fecha_registro__year=anio,
        )
        .select_related('ahorro', 'ahorro__categoria')
        .order_by('-fecha_registro')[:15]
    )

    # Normalizar a diccionarios con formato común
    movs_norm = [
        {
            'tipo':        m.tipo,
            'descripcion': m.descripcion or 'Sin descripción',
            'categoria':   m.categoria.nombre if m.categoria else 'Sin categoría',
            'fecha':       m.fecha_registro,
            'fecha_fmt':   m.fecha_registro.strftime('%d/%m/%Y'),
            'monto':       float(m.monto),
        }
        for m in movs_raw
    ]
    ahor_norm = [
        {
            'tipo':        'AHORRO',
            'descripcion': a.ahorro.descripcion or 'Aporte de ahorro',
            'categoria':   a.ahorro.categoria.nombre if a.ahorro.categoria else 'Sin categoría',
            'fecha':       a.fecha_registro,
            'fecha_fmt':   a.fecha_registro.strftime('%d/%m/%Y'),
            'monto':       float(a.aporte),
        }
        for a in ahorros_raw
    ]

    ultimos_movimientos = sorted(
        movs_norm + ahor_norm,
        key=lambda x: str(x['fecha']),
        reverse=True,
    )[:10]

    # Notificaciones — siempre las mas recientes, no dependen del mes
    notificaciones_count = Notificacion.objects.filter(
        usuario=user, leida=False,
    ).count()

    ultimas_notificaciones = (
        Notificacion.objects
        .filter(usuario=user)
        .order_by('-fecha_creacion')[:4]
    )

    return {
        'mes':                       mes,
        'anio':                      anio,
        'mes_nombre':                MESES_ES[mes],
        'es_mes_actual':             es_mes_actual,
        'total_ingresos':            total_ingresos,
        'total_egresos':             total_egresos,
        'total_ahorros':             total_ahorros,
        'utilidad':                  utilidad,
        'disponible':                disponible,
        'diferencia':                diferencia,
        'ahorro_total':              ahorro_total,
        'ahorros_mes':               ahorros_mes,
        'hay_deficit':               hay_deficit,
        'pie_data':                  pie_data,
        'pie_json':                  json.dumps(pie_data),
        'metas_ahorro_activas':      metas_ahorro_activas,
        'ultimos_movimientos':       ultimos_movimientos,
        'notificaciones_count':      notificaciones_count,
        'ultimas_notificaciones':    ultimas_notificaciones,
        'hoy':                       hoy,
    }


@login_required
def meses_disponibles(request):
    """
    Devuelve el primer mes/anio con ResumenMensual para el usuario.
    El frontend lo usa para saber hasta donde puede navegar hacia atras.
    Se llama una sola vez al cargar la pagina y se cachea en JS.
    """
    primer = (
        ResumenMensual.objects
        .filter(usuario=request.user)
        .order_by('anio', 'mes')
        .first()
    )
    hoy = date.today()
    return JsonResponse({
        'ok':          True,
        'primer_mes':  primer.mes  if primer else hoy.month,
        'primer_anio': primer.anio if primer else hoy.year,
    })


@login_required
def home_view(request):
    hoy  = date.today()
    user = request.user

    # Parsear y validar mes/anio desde query params
    try:
        mes  = int(request.GET.get('mes',  hoy.month))
        anio = int(request.GET.get('anio', hoy.year))
        if not (1 <= mes <= 12):
            mes = hoy.month
    except (ValueError, TypeError):
        mes, anio = hoy.month, hoy.year

    # No permitir navegar al futuro
    if (anio, mes) > (hoy.year, hoy.month):
        mes, anio = hoy.month, hoy.year

    ctx = _build_context(user, mes, anio)

    # Respuesta JSON para requests AJAX (navegacion sin recarga)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # ultimos_movimientos ya son dicts normalizados desde _build_context
        mov_list = [
            {
                'tipo':        m['tipo'],
                'descripcion': m['descripcion'],
                'categoria':   m['categoria'],
                'fecha':       m['fecha_fmt'],
                'monto':       str(m['monto']),
            }
            for m in ctx['ultimos_movimientos']
        ]

        notif_list = [
            {
                'titulo': n.titulo,
                'tipo':   n.tipo,
                'leida':  n.leida,
                'fecha':  n.fecha_creacion.strftime('%d/%m %H:%M'),
            }
            for n in ctx['ultimas_notificaciones']
        ]

        return JsonResponse({
            'ok':                       True,
            'mes':                      mes,
            'anio':                     anio,
            'mes_nombre':               ctx['mes_nombre'],
            'es_mes_actual':            ctx['es_mes_actual'],
            'total_ingresos':           str(ctx['total_ingresos']),
            'total_egresos':            str(ctx['total_egresos']),
            'utilidad':                 str(ctx['utilidad']),
            'disponible':               str(ctx['disponible']),
            'diferencia':               str(ctx['diferencia']),
            'ahorro_total':             str(ctx['ahorro_total']),
            'ahorros_mes':              str(ctx['ahorros_mes']),
            'hay_deficit':              ctx['hay_deficit'],
            'pie_data':                 ctx['pie_data'],
            'metas_ahorro_activas':     ctx['metas_ahorro_activas'],
            'ultimos_movimientos':      mov_list,
            'notificaciones_count':     ctx['notificaciones_count'],
            'ultimas_notificaciones':   notif_list,
        })

    return render(request, 'dashboard/home.html', ctx)


@login_required
def tendencia_mes(request):
    """
    Devuelve los totales diarios de ingresos y egresos del mes solicitado.

    Mes actual: rango completo desde dia 1 hasta hoy.
    Meses pasados: solo dias con al menos un movimiento registrado.

    Query params opcionales:
    - mes  (int, 1-12)
    - anio (int)
    """
    hoy  = date.today()
    user = request.user

    try:
        mes  = int(request.GET.get('mes',  hoy.month))
        anio = int(request.GET.get('anio', hoy.year))
    except (ValueError, TypeError):
        mes, anio = hoy.month, hoy.year

    es_mes_actual = (mes == hoy.month and anio == hoy.year)
    primer_dia    = date(anio, mes, 1)

    qs_base = Movimiento.objects.filter(
        usuario=user,
        activo=True,
        fecha_registro__month=mes,
        fecha_registro__year=anio,
    )

    def _diarios_totales(tipo):
        return {
            row['fecha']: float(row['total'])
            for row in (
                qs_base
                .filter(tipo=tipo)
                .annotate(fecha=TruncDate('fecha_registro'))
                .values('fecha')
                .annotate(total=Sum('monto'))
            )
        }

    def _diarios_por_categoria(tipo):
        from collections import defaultdict
        resultado = defaultdict(list)
        rows = (
            qs_base
            .filter(tipo=tipo)
            .annotate(fecha=TruncDate('fecha_registro'))
            .values('fecha', 'categoria__nombre')
            .annotate(total=Sum('monto'))
            .order_by('fecha', '-total')
        )
        for row in rows:
            resultado[row['fecha']].append({
                'nombre': row['categoria__nombre'] or 'Sin categoria',
                'monto':  float(row['total']),
            })
        return dict(resultado)

    ing_map     = _diarios_totales('INGRESO')
    egr_map     = _diarios_totales('EGRESO')
    ing_cat_map = _diarios_por_categoria('INGRESO')
    egr_cat_map = _diarios_por_categoria('EGRESO')

    # Ahorros diarios (solo aportes con estado APORTADO)
    qs_ahorros = AporteAhorro.objects.filter(
        ahorro__usuario=user,
        estado_ap='APORTADO',
        fecha_registro__month=mes,
        fecha_registro__year=anio,
    )

    ahor_map = {
        row['fecha_registro']: float(row['total'])
        for row in (
            qs_ahorros
            .values('fecha_registro')
            .annotate(total=Sum('aporte'))
        )
    }

    from collections import defaultdict
    ahor_cat_map = defaultdict(list)
    for row in (
        qs_ahorros
        .values('fecha_registro', 'ahorro__categoria__nombre')
        .annotate(total=Sum('aporte'))
        .order_by('fecha_registro', '-total')
    ):
        ahor_cat_map[row['fecha_registro']].append({
            'nombre': row['ahorro__categoria__nombre'] or 'Sin categoria',
            'monto':  float(row['total']),
        })
    ahor_cat_map = dict(ahor_cat_map)

    if es_mes_actual:
        total_dias = (hoy - primer_dia).days + 1
        rango = [primer_dia + timedelta(days=i) for i in range(total_dias)]
    else:
        dias_con_datos = sorted(
            set(ing_map.keys()) | set(egr_map.keys()) | set(ahor_map.keys())
        )
        rango = dias_con_datos

    if not rango:
        return JsonResponse({
            'ok': True, 'labels': [], 'ingresos': [], 'egresos': [],
            'ahorros': [], 'detalle_ing': {}, 'detalle_egr': {},
            'detalle_ahor': {}, 'total_dias': 0,
        })

    detalle_ing = {
        str(d.day): ing_cat_map.get(d, [])
        for d in rango if d in ing_cat_map
    }
    detalle_egr = {
        str(d.day): egr_cat_map.get(d, [])
        for d in rango if d in egr_cat_map
    }
    detalle_ahor = {
        str(d.day): ahor_cat_map.get(d, [])
        for d in rango if d in ahor_cat_map
    }

    return JsonResponse({
        'ok':           True,
        'labels':       [str(d.day) for d in rango],
        'ingresos':     [ing_map.get(d, 0) for d in rango],
        'egresos':      [egr_map.get(d, 0) for d in rango],
        'ahorros':      [ahor_map.get(d, 0) for d in rango],
        'detalle_ing':  detalle_ing,
        'detalle_egr':  detalle_egr,
        'detalle_ahor': detalle_ahor,
        'total_dias':   len(rango),
    })


# ═══════════════════════════════════════════════════════════════
#  EXPORTAR REPORTE EXCEL
# ═══════════════════════════════════════════════════════════════
@login_required
def exportar_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse

    hoy  = date.today()
    user = request.user
    mes  = int(request.GET.get('mes', hoy.month))
    anio = int(request.GET.get('anio', hoy.year))

    ctx = _build_context(user, mes, anio)
    movimientos = ctx['ultimos_movimientos']

    # Traer TODOS los movimientos del mes (no solo los 10 del dashboard)
    movs_all = list(
        Movimiento.objects.filter(
            usuario=user, activo=True,
            fecha_registro__month=mes, fecha_registro__year=anio,
        ).select_related('categoria').order_by('-fecha_registro')
    )
    ahor_all = list(
        AporteAhorro.objects.filter(
            ahorro__usuario=user, estado_ap='APORTADO',
            fecha_registro__month=mes, fecha_registro__year=anio,
        ).select_related('ahorro', 'ahorro__categoria').order_by('-fecha_registro')
    )

    all_items = sorted(
        [
            {
                'tipo': m.tipo, 'descripcion': m.descripcion or 'Sin descripción',
                'categoria': m.categoria.nombre if m.categoria else '—',
                'fecha': m.fecha_registro.strftime('%d/%m/%Y'),
                'sort_key': str(m.fecha_registro),
                'monto': float(m.monto),
            } for m in movs_all
        ] + [
            {
                'tipo': 'AHORRO', 'descripcion': a.ahorro.descripcion or 'Aporte',
                'categoria': a.ahorro.categoria.nombre if a.ahorro.categoria else '—',
                'fecha': a.fecha_registro.strftime('%d/%m/%Y'),
                'sort_key': str(a.fecha_registro),
                'monto': float(a.aporte),
            } for a in ahor_all
        ],
        key=lambda x: x['sort_key'],
        reverse=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f'{MESES_ES[mes]} {anio}'

    # Estilos
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Título
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f'Gastu — Reporte {MESES_ES[mes]} {anio}'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='0F172A')
    title_cell.alignment = Alignment(horizontal='center')

    # Resumen
    ws['A3'] = 'Ingresos'
    ws['B3'] = float(ctx['total_ingresos'])
    ws['B3'].number_format = '$#,##0'
    ws['C3'] = 'Egresos'
    ws['D3'] = float(ctx['total_egresos'])
    ws['D3'].number_format = '$#,##0'
    ws['A4'] = 'Ahorros'
    ws['B4'] = float(ctx['ahorros_mes'])
    ws['B4'].number_format = '$#,##0'
    ws['C4'] = 'Utilidad'
    ws['D4'] = float(ctx['utilidad'])
    ws['D4'].number_format = '$#,##0'
    for cell in ['A3', 'C3', 'A4', 'C4']:
        ws[cell].font = Font(bold=True, size=10, color='64748B')

    # Cabeceras de tabla
    headers = ['Tipo', 'Descripción', 'Categoría', 'Fecha', 'Monto']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Datos
    for i, item in enumerate(all_items, 7):
        ws.cell(row=i, column=1, value=item['tipo']).border = thin_border
        ws.cell(row=i, column=2, value=item['descripcion']).border = thin_border
        ws.cell(row=i, column=3, value=item['categoria']).border = thin_border
        ws.cell(row=i, column=4, value=item['fecha']).border = thin_border
        c = ws.cell(row=i, column=5, value=item['monto'])
        c.number_format = '$#,##0'
        c.border = thin_border
        # Color por tipo
        if item['tipo'] == 'INGRESO':
            c.font = Font(color='059669', bold=True)
        elif item['tipo'] == 'EGRESO':
            c.font = Font(color='E11D48', bold=True)
        else:
            c.font = Font(color='D97706', bold=True)

    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Gastu_Reporte_{MESES_ES[mes]}_{anio}.xlsx"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════
#  EXPORTAR REPORTE PDF
# ═══════════════════════════════════════════════════════════════
@login_required
def exportar_pdf(request):
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from django.http import HttpResponse
    import io

    hoy  = date.today()
    user = request.user
    mes  = int(request.GET.get('mes', hoy.month))
    anio = int(request.GET.get('anio', hoy.year))

    ctx = _build_context(user, mes, anio)

    # Traer TODOS los movimientos del mes
    movs_all = list(
        Movimiento.objects.filter(
            usuario=user, activo=True,
            fecha_registro__month=mes, fecha_registro__year=anio,
        ).select_related('categoria').order_by('-fecha_registro')
    )
    ahor_all = list(
        AporteAhorro.objects.filter(
            ahorro__usuario=user, estado_ap='APORTADO',
            fecha_registro__month=mes, fecha_registro__year=anio,
        ).select_related('ahorro', 'ahorro__categoria').order_by('-fecha_registro')
    )

    all_items = sorted(
        [
            {
                'tipo': m.tipo, 'descripcion': m.descripcion or 'Sin descripción',
                'categoria': m.categoria.nombre if m.categoria else '—',
                'fecha': m.fecha_registro.strftime('%d/%m/%Y'),
                'sort_key': str(m.fecha_registro),
                'monto': float(m.monto),
            } for m in movs_all
        ] + [
            {
                'tipo': 'AHORRO', 'descripcion': a.ahorro.descripcion or 'Aporte',
                'categoria': a.ahorro.categoria.nombre if a.ahorro.categoria else '—',
                'fecha': a.fecha_registro.strftime('%d/%m/%Y'),
                'sort_key': str(a.fecha_registro),
                'monto': float(a.aporte),
            } for a in ahor_all
        ],
        key=lambda x: x['sort_key'],
        reverse=True,
    )

    from django.conf import settings
    import os

    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'gastu_grafica_pdf.jpg').replace('\\', '/')

    pdf_ctx = {
        'mes_nombre':      MESES_ES[mes],
        'anio':            anio,
        'total_ingresos':  float(ctx['total_ingresos']),
        'total_egresos':   float(ctx['total_egresos']),
        'ahorros_mes':     float(ctx['ahorros_mes']),
        'utilidad':        float(ctx['utilidad']),
        'disponible':      float(ctx['disponible']),
        'movimientos':     all_items,
        'usuario':         user.username,
        'fecha_generacion': hoy.strftime('%d/%m/%Y %H:%M'),
        'logo_path':       logo_path,
    }

    html_string = render_to_string('dashboard/reporte_pdf.html', pdf_ctx)

    buffer = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=buffer, encoding='utf-8')

    if pisa_status.err:
        return HttpResponse('Error generando el PDF', status=500)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Gastu_Reporte_{MESES_ES[mes]}_{anio}.pdf"'
    return response