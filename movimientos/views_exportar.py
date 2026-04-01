"""
Vistas de exportación de movimientos.
Soporta CSV, Excel (openpyxl) y PDF (xhtml2pdf).

Parámetros GET comunes:
- tipo        : INGRESO | EGRESO | AMBOS
- fecha_desde : YYYY-MM-DD
- fecha_hasta : YYYY-MM-DD
- categorias  : ids separados por coma, o vacío = todas
"""
import csv
import io
from datetime import date, datetime
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from categorias.models import Categoria
from .models import Movimiento


# ── Helpers ──────────────────────────────────────────────────────────────────

def _parse_fecha(valor, fallback):
    """Convierte string YYYY-MM-DD a date. Devuelve fallback si falla."""
    try:
        return datetime.strptime(valor, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return fallback


def _build_qs(request):
    """
    Construye el queryset de movimientos filtrado según los parámetros GET.
    Siempre filtra por request.user.
    """
    hoy = timezone.now().date()
    tipo        = request.GET.get('tipo', 'AMBOS').upper()
    fecha_desde = _parse_fecha(request.GET.get('fecha_desde'), date(hoy.year, hoy.month, 1))
    fecha_hasta = _parse_fecha(request.GET.get('fecha_hasta'), hoy)
    cat_ids_raw = request.GET.get('categorias', '')

    qs = Movimiento.objects.filter(
        usuario=request.user,
        activo=True,
        fecha_registro__date__gte=fecha_desde,
        fecha_registro__date__lte=fecha_hasta,
    ).select_related('categoria').order_by('fecha_registro')

    if tipo in ('INGRESO', 'EGRESO'):
        qs = qs.filter(tipo=tipo)

    if cat_ids_raw:
        try:
            cat_ids = [int(x) for x in cat_ids_raw.split(',') if x.strip()]
            if cat_ids:
                qs = qs.filter(categoria_id__in=cat_ids)
        except ValueError:
            pass

    return qs, fecha_desde, fecha_hasta, tipo


def _nombre_archivo(tipo, fecha_desde, fecha_hasta, extension):
    """Genera el nombre del archivo de descarga."""
    tipo_str = tipo.lower() if tipo != 'AMBOS' else 'movimientos'
    return f"gastuapp_{tipo_str}_{fecha_desde}_{fecha_hasta}.{extension}"


def _resumen(qs):
    """Calcula totales de ingresos y egresos sobre el queryset."""
    total_ingresos = qs.filter(tipo='INGRESO').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_egresos  = qs.filter(tipo='EGRESO').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    balance        = total_ingresos - total_egresos
    return total_ingresos, total_egresos, balance


# ── CSV ───────────────────────────────────────────────────────────────────────

@login_required
def exportar_csv(request):
    """Exporta movimientos a CSV."""
    qs, fecha_desde, fecha_hasta, tipo = _build_qs(request)
    nombre = _nombre_archivo(tipo, fecha_desde, fecha_hasta, 'csv')

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'

    writer = csv.writer(response)
    writer.writerow(['Fecha', 'Tipo', 'Categoria', 'Descripcion', 'Monto'])

    for m in qs:
        writer.writerow([
            m.fecha_registro.date().strftime('%d/%m/%Y'),
            m.get_tipo_display(),
            m.categoria.nombre,
            m.descripcion or '',
            str(m.monto),
        ])

    total_ingresos, total_egresos, balance = _resumen(qs)
    writer.writerow([])
    writer.writerow(['', '', '', 'Total ingresos', str(total_ingresos)])
    writer.writerow(['', '', '', 'Total egresos',  str(total_egresos)])
    writer.writerow(['', '', '', 'Balance',         str(balance)])

    return response


# ── Excel ─────────────────────────────────────────────────────────────────────

@login_required
def exportar_excel(request):
    """Exporta movimientos a Excel usando openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl no está instalado. Ejecuta: pip install openpyxl', status=500)

    qs, fecha_desde, fecha_hasta, tipo = _build_qs(request)
    nombre = _nombre_archivo(tipo, fecha_desde, fecha_hasta, 'xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Movimientos'

    # Paleta
    COLOR_HEADER  = '0F172A'
    COLOR_INGRESO = 'D1FAE5'
    COLOR_EGRESO  = 'FFEDD5'
    COLOR_TOTAL   = 'F1F5F9'

    thin = Side(style='thin', color='E2E8F0')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells('A1:E1')
    titulo = ws['A1']
    titulo.value = f'GastuApp — Reporte de movimientos ({fecha_desde.strftime("%d/%m/%Y")} al {fecha_hasta.strftime("%d/%m/%Y")})'
    titulo.font      = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    titulo.fill      = PatternFill('solid', fgColor=COLOR_HEADER)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Encabezados
    encabezados = ['Fecha', 'Tipo', 'Categoría', 'Descripción', 'Monto']
    for col, titulo_col in enumerate(encabezados, 1):
        celda = ws.cell(row=2, column=col, value=titulo_col)
        celda.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        celda.fill      = PatternFill('solid', fgColor='334155')
        celda.alignment = Alignment(horizontal='center', vertical='center')
        celda.border    = borde
    ws.row_dimensions[2].height = 20

    # Filas de datos
    for fila, m in enumerate(qs, 3):
        color_fila = COLOR_INGRESO if m.tipo == 'INGRESO' else COLOR_EGRESO
        valores = [
            m.fecha_registro.date().strftime('%d/%m/%Y'),
            m.get_tipo_display(),
            m.categoria.nombre,
            m.descripcion or '',
            float(m.monto),
        ]
        for col, val in enumerate(valores, 1):
            celda = ws.cell(row=fila, column=col, value=val)
            celda.fill   = PatternFill('solid', fgColor=color_fila)
            celda.border = borde
            celda.font   = Font(name='Calibri', size=10)
            if col == 5:
                celda.number_format = '#,##0.00'
                celda.alignment     = Alignment(horizontal='right')

    # Totales
    ultima_fila = qs.count() + 3
    total_ingresos, total_egresos, balance = _resumen(qs)

    for offset, (label, valor) in enumerate([
        ('Total ingresos', float(total_ingresos)),
        ('Total egresos',  float(total_egresos)),
        ('Balance',        float(balance)),
    ]):
        fila_tot = ultima_fila + offset + 1
        ws.merge_cells(f'A{fila_tot}:D{fila_tot}')
        etiqueta = ws.cell(row=fila_tot, column=1, value=label)
        etiqueta.font      = Font(name='Calibri', bold=True, size=10)
        etiqueta.fill      = PatternFill('solid', fgColor=COLOR_TOTAL)
        etiqueta.alignment = Alignment(horizontal='right')
        etiqueta.border    = borde

        monto_cell = ws.cell(row=fila_tot, column=5, value=valor)
        monto_cell.font          = Font(name='Calibri', bold=True, size=10)
        monto_cell.fill          = PatternFill('solid', fgColor=COLOR_TOTAL)
        monto_cell.number_format = '#,##0.00'
        monto_cell.alignment     = Alignment(horizontal='right')
        monto_cell.border        = borde

    # Anchos de columna
    anchos = [14, 12, 22, 38, 16]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


# ── PDF ───────────────────────────────────────────────────────────────────────

@login_required
def exportar_pdf(request):
    """Exporta movimientos a PDF usando xhtml2pdf con plantilla corporativa Gastu."""
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return HttpResponse('xhtml2pdf no está instalado. Ejecuta: pip install xhtml2pdf', status=500)

    qs, fecha_desde, fecha_hasta, tipo = _build_qs(request)
    nombre = _nombre_archivo(tipo, fecha_desde, fecha_hasta, 'pdf')
    total_ingresos, total_egresos, balance = _resumen(qs)

    tipo_str = {'INGRESO': 'Ingresos', 'EGRESO': 'Egresos', 'AMBOS': 'Ingresos y Egresos'}.get(tipo, tipo)

    # Ruta absoluta del logo para xhtml2pdf (requiere path del sistema de archivos)
    logo_path = finders.find('img/gastu_grafica_pdf.jpg')
    if logo_path:
        logo_path = logo_path.replace('\\', '/')

    contexto = {
        'logo_path':        logo_path,
        'tipo_str':         tipo_str,
        'fecha_desde':      fecha_desde.strftime('%d/%m/%Y'),
        'fecha_hasta':      fecha_hasta.strftime('%d/%m/%Y'),
        'usuario':          request.user.get_full_name() or request.user.username,
        'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'movimientos':      list(qs),
        'total_ingresos':   total_ingresos,
        'total_egresos':    total_egresos,
        'balance':          balance,
        'cantidad':         qs.count(),
    }

    html_string = render_to_string('movimientos/reporte_pdf.html', contexto)

    buffer = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=buffer)

    if pisa_status.err:
        return HttpResponse('Error al generar el PDF. Intenta nuevamente.', status=500)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response